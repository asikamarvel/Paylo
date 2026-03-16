"""Report generation utilities."""

from datetime import datetime
from pathlib import Path
from decimal import Decimal

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from .calculator_v2 import PayrollBreakdown
from .currency import CurrencyConverter, MockCurrencyConverter


class PayrollReport:
    
    def __init__(self, output_dir: str = None):
        if output_dir is None:
            output_dir = Path(__file__).parent.parent / "output"
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_excel_report(self, payroll_data: list, period: str, filename: str = None, 
                              currency_converter: CurrencyConverter = None) -> str:
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"payroll_report_{period.replace(' ', '_')}_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        wb = openpyxl.Workbook()
        self._create_summary_sheet(wb, payroll_data, period)
        self._create_detail_sheet(wb, payroll_data, period)
        self._create_hours_sheet(wb, payroll_data, period)
        
        wb.save(filepath)
        return str(filepath)
    
    def _create_summary_sheet(self, wb, payroll_data: list, period: str):
        ws = wb.active
        ws.title = "Summary"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        ws["A1"] = f"Payroll Summary - {period}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")
        
        headers = ["Name", "Cadre", "Type", "Gross Pay", "Deductions", "Net Pay"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, payroll in enumerate(payroll_data, 4):
            symbol = "₦" if payroll.currency == "NGN" else "$"
            
            ws.cell(row=row_idx, column=1, value=payroll.staff.name).border = border
            ws.cell(row=row_idx, column=2, value=payroll.staff.cadre.replace("_", " ").title()).border = border
            ws.cell(row=row_idx, column=3, value=payroll.staff.contractor_type.title()).border = border
            ws.cell(row=row_idx, column=4, value=f"{symbol}{float(payroll.gross_compensation):,.2f}").border = border
            ws.cell(row=row_idx, column=5, value=f"{symbol}{float(payroll.total_deductions):,.2f}").border = border
            ws.cell(row=row_idx, column=6, value=f"{symbol}{float(payroll.net_compensation):,.2f}").border = border
        
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_detail_sheet(self, wb, payroll_data: list, period: str):
        ws = wb.create_sheet("Detailed Breakdown")
        
        headers = ["Name", "Base Payment", "Local Billable", "Regional Billable", 
                   "Gross", "WHT", "HMO", "Net", "Currency"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, payroll in enumerate(payroll_data, 2):
            ws.cell(row=row_idx, column=1, value=payroll.staff.name)
            ws.cell(row=row_idx, column=2, value=float(payroll.base_payment))
            ws.cell(row=row_idx, column=3, value=float(payroll.local_billable_fees))
            ws.cell(row=row_idx, column=4, value=float(payroll.regional_billable_fees))
            ws.cell(row=row_idx, column=5, value=float(payroll.gross_compensation))
            ws.cell(row=row_idx, column=6, value=float(payroll.withholding_tax))
            ws.cell(row=row_idx, column=7, value=float(payroll.hmo_deduction))
            ws.cell(row=row_idx, column=8, value=float(payroll.net_compensation))
            ws.cell(row=row_idx, column=9, value=payroll.currency)
    
    def _create_hours_sheet(self, wb, payroll_data: list, period: str):
        ws = wb.create_sheet("Hours Breakdown")
        
        headers = ["Name", "Total Hours", "Billable", "Non-Billable", "Extra Billable"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, payroll in enumerate(payroll_data, 2):
            ws.cell(row=row_idx, column=1, value=payroll.staff.name)
            ws.cell(row=row_idx, column=2, value=payroll.hours.total_hours)
            ws.cell(row=row_idx, column=3, value=payroll.hours.total_billable_hours)
            ws.cell(row=row_idx, column=4, value=payroll.hours.non_billable_hours)
            ws.cell(row=row_idx, column=5, value=payroll.hours.extra_billable_hours)


def generate_report(payroll_data: list, period: str, output_dir: str = None, 
                    currency_converter: CurrencyConverter = None) -> str:
    report = PayrollReport(output_dir=output_dir)
    return report.generate_excel_report(payroll_data, period, currency_converter=currency_converter)
